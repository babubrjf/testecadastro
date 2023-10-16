import openpyxl
import time
import os
from tabulate import tabulate
import pandas as pd

# Função para cadastrar um novo cliente
def addCad():
    os.system('cls')
    workbook = openpyxl.load_workbook(r"\\servidor\Turmas\dev-sis-manha\0000999171\cadastro.xlsx")
    sheet = workbook["Clientes"]
    nome = input("Digite o nome do cliente: ")
    cpf = input("Digite o CPF do cliente: ")
    tel  = input("Digite o telefone do cliente: ")
    email = input("Digite o email do cliente: ")
    endereco = input("Digite o endereço do cliente: ")
    data = input("Digite a data de nascimento do cliente: ")
    sheet.append([nome, cpf, tel, email, endereco, data])
    workbook.save(r"\\servidor\Turmas\dev-sis-manha\0000999171\cadastro.xlsx")
    print("Cliente Cadastrado Com Sucesso!")
    time.sleep(3)
    os.system('cls')

# Função para atualizar informações de um cliente pelo CPF
def attCad():
    os.system('cls')
    workbook = openpyxl.load_workbook(r"\\servidor\Turmas\dev-sis-manha\0000999171\cadastro.xlsx")
    sheet = workbook["Clientes"]
    cpf = int(input("Digite o CPF do cliente para atualizar seus dados:"))
    userId = None
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            if (int(cell.value) == cpf):
                userId = cell
                break
    if userId == None:
        print("Cliente não encontrado!")
        time.sleep(3)
        os.system('cls')
    else:
        nomen = input("Digite o novo nome: ")
        cpfn = input("Digite o novo CPF: ")
        teln  = input("Digite o novo telefone: ")
        emailn = input("Digite o novo email: ")
        enderecon = input("Digite o novo endereço: ")
        datan = input("Digite a nova data de nascimento: ")
        dadosA = ["", nomen, cpfn, teln, emailn, enderecon, datan]
        for i in range(1, 7):
            sheet.cell(userId.row, i, dadosA[i])
        workbook.save(r"\\servidor\Turmas\dev-sis-manha\0000999171\cadastro.xlsx")
        print("Cliente Atualizado Com Sucesso!")
        time.sleep(3)
        os.system('cls')


# Função para listar todos os clientes
def listaCad():
    os.system('cls')
    clientes = []
    op = "1"
    while op == "1":
        workbook = openpyxl.load_workbook(r"\\servidor\Turmas\dev-sis-manha\0000999171\cadastro.xlsx")
        sheet = workbook["Clientes"]
        for row in sheet.iter_rows(min_row=1, values_only=True):
            clientes.append(row)
        print(tabulate(clientes, headers=[], tablefmt= "fancy_grid"))
        op = input("Pressione ENTER p/ voltar ao menu: ")
        os.system('cls')
        

# Função para excluir um cliente pelo CPF
def delCad():
    os.system('cls')
    workbook = openpyxl.load_workbook(r"\\servidor\Turmas\dev-sis-manha\0000999171\cadastro.xlsx")
    sheet = workbook["Clientes"]
    cpf = int(input("Digite o CPF do cliente para excluir o cadastro:"))
    userId = None
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            if (int(cell.value) == cpf):
                userId = cell
                break
    if userId == None:
        print("Cliente não encontrado no sistema!")
        time.sleep(3)
        os.system('cls')
    else:
        sheet.delete_rows(cell.row)
        workbook.save(r"\\servidor\Turmas\dev-sis-manha\0000999171\cadastro.xlsx")
        print("Cliente excluído com sucesso!")
        time.sleep(3)
        os.system('cls')

# Menu:
menu=1
os.system('cls')
while menu:
    print("1. Cadastrar novo cliente")
    print("2. Atualizar dados do cliente")
    print("3. Listar clientes cadastrados")
    print("4. Excluir cadastros ")
    print("0. Sair")
    menu = int(input("Opção: "))

    if(menu==1):
        addCad()
    if(menu==2):
        attCad()
    if(menu==3):
        listaCad()
    if(menu==4):
        delCad()
    if(menu==0):
        raise SystemExit